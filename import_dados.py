import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import pandas as pd
import mysql.connector
from mysql.connector import errorcode
from openpyxl import load_workbook
import os  # Para verificar se o arquivo existe
import math  # Para verificar valores NaN

def select_sheet():
    file_path = file_path_entry.get()
    if file_path and os.path.exists(file_path) and file_path.endswith('.xlsx'):
        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            sheet_selector['values'] = sheet_names
            sheet_selector.current(0)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar o arquivo: {e}")
    else:
        sheet_selector['values'] = ['N/A']
        sheet_selector.current(0)
        messagebox.showwarning("Aviso", "Selecione um arquivo .xlsx válido.")

def upload_data():
    file_path = file_path_entry.get()
    file_type = file_type_entry.get().strip().lower()
    sheet_name = sheet_selector.get()
    db_name = db_name_entry.get().strip()
    table_name = table_name_entry.get().strip()

    if not file_path or not os.path.exists(file_path):
        messagebox.showerror("Erro", "Caminho do arquivo inválido ou arquivo não encontrado.")
        return

    if file_type not in ["xlsx", "txt"]:
        messagebox.showerror("Erro", "Tipo de arquivo inválido. Use 'xlsx' ou 'txt'.")
        return

    if not db_name or not table_name:
        messagebox.showerror("Erro", "Nome do banco de dados e da tabela são obrigatórios.")
        return

    cursor = None
    cnx = None
    try:
        # Conectar ao MySQL
        cnx = mysql.connector.connect(
            user='root',
            password='@Kaclju2125.',
            host='localhost',
            port=3306
        )
        cursor = cnx.cursor()

        # Verificar e criar banco de dados se não existir
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {db_name}")
        cursor.execute(f"USE {db_name}")

        # Verificar se a tabela já existe
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        table_exists = cursor.fetchone()

        # Se a tabela existir, excluí-la
        if table_exists:
            cursor.execute(f"DROP TABLE `{table_name}`")
            messagebox.showinfo("Info", f"Tabela '{table_name}' existente foi excluída para substituição.")

        # Ler dados do arquivo
        if file_type == "xlsx":
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        elif file_type == "txt":
            df = pd.read_csv(file_path, delimiter='\t')

        # Verificar se o DataFrame está vazio
        if df.empty:
            messagebox.showwarning("Aviso", "O arquivo está vazio ou não contém dados.")
            return

        # Adicionar coluna 'id' se não existir
        if 'id' not in df.columns:
            df.insert(0, 'id', None)  # Adiciona a coluna 'id' no início do DataFrame

        # Definir tipos de colunas dinamicamente
        columns_with_types = []
        for col in df.columns:
            if col == 'id':
                columns_with_types.append('`id` INT AUTO_INCREMENT PRIMARY KEY')  # Coluna 'id' como chave primária
            else:
                max_length = df[col].apply(lambda x: len(str(x)) if x else 0).max()
                if max_length <= 255:
                    columns_with_types.append(f'`{col}` VARCHAR(255)')
                else:
                    columns_with_types.append(f'`{col}` TEXT')

        # Criar a tabela com a nova estrutura
        cursor.execute(f"""
            CREATE TABLE `{table_name}` (
                {', '.join(columns_with_types)}
            )
        """)

        # Inserir ou substituir dados na tabela
        for _, row in df.iterrows():
            # Substituir NaN por None (NULL no MySQL)
            row = [None if pd.isna(value) else value for value in row]
            # Remover o valor da coluna 'id' para permitir auto-incremento
            if 'id' in df.columns:
                row = row[1:]  # Remove o valor da coluna 'id'
            placeholders = ', '.join(['%s'] * len(row))
            columns = ', '.join([f'`{col}`' for col in df.columns if col != 'id'])
            values = tuple(row)
            cursor.execute(f"""
                INSERT INTO `{table_name}` ({columns}) 
                VALUES ({placeholders})
            """, values)

        cnx.commit()
        messagebox.showinfo("Sucesso", "Dados enviados e tabela atualizada com sucesso.")
    except mysql.connector.Error as err:
        messagebox.showerror("Erro", f"Erro no MySQL: {err}")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro inesperado: {e}")
    finally:
        if cursor:
            cursor.close()
        if cnx:
            cnx.close()

# Configurações da interface gráfica
root = tk.Tk()
root.title("Upload de Dados para MySQL")

tk.Label(root, text="Caminho do Arquivo:").grid(row=0, column=0)
file_path_entry = tk.Entry(root, width=50)
file_path_entry.grid(row=0, column=1)
tk.Button(root, text="Browse", command=lambda: file_path_entry.insert(0, filedialog.askopenfilename())).grid(row=0, column=2)
tk.Button(root, text="Selecionar Aba", command=select_sheet).grid(row=0, column=3)

tk.Label(root, text="Tipo de Arquivo (xlsx ou txt):").grid(row=1, column=0)
file_type_entry = tk.Entry(root)
file_type_entry.grid(row=1, column=1)

tk.Label(root, text="Nome da Aba:").grid(row=2, column=0)
sheet_selector = ttk.Combobox(root)
sheet_selector.grid(row=2, column=1)

tk.Label(root, text="Nome do Banco de Dados:").grid(row=3, column=0)
db_name_entry = tk.Entry(root)
db_name_entry.grid(row=3, column=1)

tk.Label(root, text="Nome da Tabela:").grid(row=4, column=0)
table_name_entry = tk.Entry(root)
table_name_entry.grid(row=4, column=1)

tk.Button(root, text="Upload", command=upload_data).grid(row=5, column=0, columnspan=3)

root.mainloop()