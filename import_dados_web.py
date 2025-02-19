import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import mysql.connector
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials

def select_sheet():
    file_path = file_path_entry.get()
    if file_type_entry.get() == "google_sheet":
        sheet_selector['values'] = [file_path]
        sheet_selector.current(0)
    elif file_path.endswith('.xlsx'):
        workbook = load_workbook(file_path, read_only=True)
        sheet_names = workbook.sheetnames
        sheet_selector['values'] = sheet_names
        sheet_selector.current(0)
    else:
        sheet_selector['values'] = ['N/A']
        sheet_selector.current(0)

def get_google_sheet_data(credentials_path, sheet_url, sheet_name):
    try:
        scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        credentials = Credentials.from_service_account_file(credentials_path, scopes=scope)
        client = gspread.authorize(credentials)
        spreadsheet_id = sheet_url.split('/d/')[1].split('/')[0]
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet(sheet_name)
        data = worksheet.get_all_records()
        df = pd.DataFrame(data)
        
        # Substituir valores infinitos ou NaN
        df.replace([float('inf'), -float('inf')], None, inplace=True)
        df.fillna('', inplace=True)  # Preencher NaN com string vazia ou valor adequado

        return df
    except gspread.exceptions.APIError as api_err:
        print(f"Erro de API: {api_err}")
        raise
    except Exception as e:
        print(f"Ocorreu um erro: {e}")
        raise

def map_dtype(dtype):
    if dtype == 'object':
        return 'TEXT(255)'
    elif dtype.startswith('int'):
        return 'INT'
    elif dtype.startswith('float'):
        return 'FLOAT'
    else:
        return 'TEXT(255)'

def upload_data():
    credentials_path = json_path_entry.get()
    sheet_url = url_entry.get()
    sheet_name = sheet_selector.get()
    db_name = db_name_entry.get()
    table_name = table_name_entry.get()

    cursor = None
    cnx = None
    try:
        cnx = mysql.connector.connect(
            user='root',
            password='@Kaclju2125.',
            host='localhost',
            port=3306
        )
        cursor = cnx.cursor()
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {db_name}")
        cursor.execute(f"USE {db_name}")
        df = get_google_sheet_data(credentials_path, sheet_url, sheet_name)
        
        # Adicionar uma coluna 'ID' se não existir
        if 'ID' not in df.columns:
            df.insert(0, 'ID', range(1, len(df) + 1))  # Cria uma coluna 'ID' com valores únicos

        # Confirmar que não há valores infinitos ou NaN nos dados
        df.replace([float('inf'), -float('inf')], None, inplace=True)
        df.fillna('', inplace=True)  # Preencher NaN com string vazia ou valor adequado

        # Remover linhas duplicadas
        df.drop_duplicates(subset=['ID'], inplace=True)  # Remove duplicatas com base na coluna 'ID'
        
        print("Número de linhas no dataframe:", len(df))
        print(df.head())

        # Criar a tabela se não existir
        columns_with_types = ', '.join([f'{col} {map_dtype(df[col].dtype.name)}' for col in df.columns])
        cursor.execute(f"""
            CREATE TABLE IF NOT EXISTS {table_name} (
                {columns_with_types},
                PRIMARY KEY (ID)
            )
        """)
        
        # Inserir ou atualizar dados em lotes
        batch_size = 10000  # Tamanho do lote para inserção
        for i in range(0, len(df), batch_size):
            batch = df[i:i + batch_size]
            values = [tuple(row) for row in batch.to_numpy()]
            placeholders = ', '.join(['%s'] * len(df.columns))
            update_columns = ', '.join([f'{col}=VALUES({col})' for col in df.columns])
            cursor.executemany(f"""
                INSERT INTO {table_name} ({', '.join(df.columns)}) 
                VALUES ({placeholders})
                ON DUPLICATE KEY UPDATE
                {update_columns}
            """, values)
            cnx.commit()
            print(f"Linhas {i + 1} a {i + len(batch)} inseridas/atualizadas com sucesso.")
        
        messagebox.showinfo("Sucesso", "Dados enviados e atualizados com sucesso")
    except mysql.connector.Error as err:
        messagebox.showerror("Erro de MySQL", str(err))
        print(f"Erro de MySQL: {err}")
    except gspread.exceptions.APIError as api_err:
        messagebox.showerror("Erro de API", str(api_err))
        print(f"Erro de API: {api_err}")
    except Exception as e:
        messagebox.showerror("Erro", str(e))
        print(f"Erro: {e}")
    finally:
        if cursor:
            cursor.close()
        if cnx:
            cnx.close()

root = tk.Tk()
root.title("Upload de Dados para MySQL")

tk.Label(root, text="Caminho para o arquivo de credenciais JSON:").grid(row=0, column=0)
json_path_entry = tk.Entry(root)
json_path_entry.grid(row=0, column=1)
tk.Button(root, text="Browse", command=lambda: json_path_entry.insert(0, filedialog.askopenfilename())).grid(row=0, column=2)

tk.Label(root, text="URL do Google Sheet:").grid(row=1, column=0)
url_entry = tk.Entry(root)
url_entry.grid(row=1, column=1)

tk.Label(root, text="Nome da Aba:").grid(row=2, column=0)
sheet_selector = ttk.Combobox(root)
sheet_selector.grid(row=2, column=1)

tk.Label(root, text="Nome do Banco de Dados:").grid(row=3, column=0)
db_name_entry = tk.Entry(root)
db_name_entry.grid(row=3, column=1)

tk.Label(root, text="Nome da Tabela:").grid(row=4, column=0)
table_name_entry = tk.Entry(root)
table_name_entry.grid(row=4, column=1)

tk.Button(root, text="Upload", command=upload_data).grid(row=5, column=0, columnspan=3)

root.mainloop()